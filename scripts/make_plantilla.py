# Genera public/plantilla-viajes.xlsx (la plantilla descargable del modal de
# carga por Excel). Sigue el modelo "INCOMING HUB | PLANTILLA SERVICIOS"
# provisto por el cliente:
#   - Hoja "Viajes": título en la fila 2 (A2:L2 combinadas), encabezados en la
#     fila 4 (FECHA, HORA, CATEGORÍA, PAX, TELÉFONO, TIPO, ORIGEN, DESTINO,
#     VUELO, OBSERVACIONES, DESTINO 2, DESTINO 3) y datos desde la fila 5.
#   - FECHA es una celda de fecha real (se muestra DD/MM/AAAA; el parser lee el
#     valor, no el texto, así que el formato de visualización no importa).
#   - HORA es una celda de hora real (HH:MM).
#   - Una fila por VIAJE. Tramos adicionales del mismo viaje se cargan en las
#     columnas "DESTINO 2" y "DESTINO 3" (no se repiten filas).
#   - TELÉFONO: uno por pasajero, mismo orden, separados con " | ".
#   - TIPO y CATEGORÍA con desplegable (validación de datos) desde la hoja LOV
#     (oculta), aplicado a toda la grilla (filas 5 a 203).
#   - Filas de ejemplo (una por tipo de servicio) y el resto de la grilla vacía
#     pero ya con bordes, para que el cliente solo complete.
#   - Hoja "Instrucciones" con la explicación de cada columna.
#
# NOTA: el parseo del Excel ocurre en el front (src/lib/excelParse.ts). Si se
# cambian columnas o los valores de Tipo/Categoría, hay que actualizar el mapa
# COLS de ese módulo (y los tests de src/lib/excelParse.test.ts).
#
# Uso:  python scripts/make_plantilla.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, time, timedelta

TITLE = "INCOMING HUB | PLANTILLA SERVICIOS"
TITLE_ROW = 2
HEADER_ROW = 4
FIRST_DATA_ROW = 5
LAST_GRID_ROW = 64      # filas con bordes listas para completar
LAST_DV_ROW = 203       # hasta dónde llegan los desplegables

wb = Workbook()
ws = wb.active
ws.title = "Viajes"

headers = [
    "FECHA",          # A
    "HORA",           # B
    "CATEGORÍA",      # C
    "PAX",            # D
    "TELÉFONO",       # E
    "TIPO",           # F
    "ORIGEN",         # G
    "DESTINO",        # H
    "VUELO",          # I
    "OBSERVACIONES",  # J
    "DESTINO 2",      # K
    "DESTINO 3",      # L
]
NCOLS = len(headers)

# Listas de valores (deben coincidir con la hoja LOV de abajo).
TIPOS = ["Llegada (in)", "Salida (out)", "Hs Disposición", "Otro"]
CATEGORIAS = ["Auto Std", "Ejecutivo", "MB", "Vito"]

F = date.today() + timedelta(days=1)

# Filas de ejemplo. Una fila = un viaje. Columnas:
# FECHA, HORA, CATEGORÍA, PAX, TELÉFONO, TIPO, ORIGEN, DESTINO,
# VUELO, OBSERVACIONES, DESTINO 2, DESTINO 3
rows = [
    # Llegada con vuelo, 1 pasajero
    [F, time(7, 30), "Ejecutivo", "JUAN LÓPEZ", "+54 9 11 5555-1234",
     "Llegada (in)", "AEROPUERTO EZEIZA (EZE)", "725 CONTINENTAL", "AA 909", "", "", ""],

    # Salida con vuelo, 2 pasajeros (separados con  |  ) y 2 teléfonos alineados
    [F + timedelta(days=1), time(10, 0), "Auto Std", "MARÍA CÁCERES | N. FABBRI",
     "+54 9 11 5555-1234 | +54 9 11 6033-2210",
     "Salida (out)", "SANTOS DUMONT 3429, CABA", "AEROPARQUE (AEP)", "AR1256", "OFICINA", "", ""],

    # Salida con vuelo, categoría Vito
    [F + timedelta(days=2), time(14, 0), "Vito", "LUCAS GARIBALDI", "+54 9 387 555-1456",
     "Salida (out)", "MARRIOT PLAZA", "AEROPUERTO EZEIZA (EZE)", "AR1456", "", "", ""],

    # Traslado de varios tramos: Origen -> Destino -> Destino 2 -> Destino 3
    [F + timedelta(days=3), time(14, 0), "Ejecutivo", "ELISA BARÓ", "+54 9 11 5555-4302",
     "Otro", "HOTEL FAENA", "HOTEL FAENA", "", "CITY TOUR", "PUERTO MADERO", "HOTEL FAENA"],

    # Horas a disposición (sin vuelo), categoría MB
    [F + timedelta(days=4), time(10, 0), "MB", "F. ACOSTA", "+54 9 11 5555-7788",
     "Hs Disposición", "HOTEL ALVEAR", "HOTEL ALVEAR", "", "4 HS A DISPOSICIÓN", "", ""],
]

# --- estilos ---
HEADER_BG = "1F2937"
GRID_COLOR = "D1D5DB"

title_font = Font(name="Arial", size=12, bold=True, color=HEADER_BG)
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color=HEADER_BG)
thin = Side(border_style="thin", color=GRID_COLOR)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
body_font = Font(name="Arial", size=11)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Título (fila 2, combinada a lo ancho de la tabla)
ws.cell(row=TITLE_ROW, column=1, value=TITLE).font = title_font
ws.cell(row=TITLE_ROW, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=NCOLS)
ws.row_dimensions[TITLE_ROW].height = 32

# Encabezados (fila 4)
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
ws.row_dimensions[HEADER_ROW].height = 26

# Columnas centradas (fecha/hora) vs. alineadas a la izquierda (texto)
CENTERED_COLS = {1, 2}

# Grilla: filas de ejemplo + filas vacías ya formateadas
for i in range(FIRST_DATA_ROW, LAST_GRID_ROW + 1):
    data = rows[i - FIRST_DATA_ROW] if i - FIRST_DATA_ROW < len(rows) else None
    for col_idx in range(1, NCOLS + 1):
        cell = ws.cell(row=i, column=col_idx)
        if data is not None and data[col_idx - 1] != "":
            cell.value = data[col_idx - 1]
        cell.border = border
        cell.font = body_font
        cell.alignment = center if col_idx in CENTERED_COLS else left
    # Celdas de fecha y hora reales, con formato de visualización fijo.
    ws.cell(row=i, column=1).number_format = "DD/MM/YYYY"
    ws.cell(row=i, column=2).number_format = "HH:MM"
    ws.row_dimensions[i].height = 27

# Anchos
widths = {
    "A": 12, "B": 8, "C": 12, "D": 28, "E": 18, "F": 16,
    "G": 30, "H": 30, "I": 9, "J": 26, "K": 24, "L": 24,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = f"A{FIRST_DATA_ROW}"
ws.sheet_view.showGridLines = False

# --- Hoja LOV (listas para los desplegables), oculta ---
lov = wb.create_sheet("LOV")
lov["A1"] = "Tipo"
lov["B1"] = "Categoria"
for c in ("A1", "B1"):
    lov[c].font = header_font
    lov[c].fill = header_fill
for i, v in enumerate(TIPOS, start=2):
    lov.cell(row=i, column=1, value=v).font = body_font
for i, v in enumerate(CATEGORIAS, start=2):
    lov.cell(row=i, column=2, value=v).font = body_font
lov.column_dimensions["A"].width = 18
lov.column_dimensions["B"].width = 14
lov.sheet_state = "hidden"

# Validación de datos (desplegables) sobre toda la columna útil, sin huecos
dv_tipo = DataValidation(
    type="list", formula1=f"=LOV!$A$2:$A${len(TIPOS) + 1}", allow_blank=True)
dv_tipo.error = "Elegí un Tipo de la lista."
dv_tipo.prompt = "Elegí: Llegada (in) / Salida (out) / Hs Disposición / Otro"
ws.add_data_validation(dv_tipo)
dv_tipo.add(f"F{FIRST_DATA_ROW}:F{LAST_DV_ROW}")

dv_cat = DataValidation(
    type="list", formula1=f"=LOV!$B$2:$B${len(CATEGORIAS) + 1}", allow_blank=True)
dv_cat.error = "Elegí una Categoría de la lista."
dv_cat.prompt = "Elegí: Auto Std / Ejecutivo / MB / Vito"
ws.add_data_validation(dv_cat)
dv_cat.add(f"C{FIRST_DATA_ROW}:C{LAST_DV_ROW}")

# --- Hoja de instrucciones ---
ws2 = wb.create_sheet("Instrucciones")
ws2.column_dimensions["A"].width = 27
ws2.column_dimensions["B"].width = 110

bold = Font(name="Arial", size=11, bold=True)
norm = Font(name="Arial", size=11)

ws2["A1"] = TITLE
ws2["A1"].font = title_font
ws2.merge_cells("A1:B1")
ws2.row_dimensions[1].height = 31

instrucciones = [
    ("HOJA VIAJES", "Una fila por VIAJE. Para tramos adicionales usá las columnas DESTINO 2 y DESTINO 3. "
                    "Las filas de ejemplo se pueden borrar o pisar."),
    ("FECHA", "Formato día/mes/año (ej: 10/09/2026)."),
    ("HORA", "Formato HH:MM en 24 horas (ej: 07:30)."),
    ("CATEGORÍA", "Elegir del desplegable: Auto Std / Ejecutivo / MB / Vito."),
    ("PAX", "Nombre del pasajero. Varios separados con  |  (pipe). Máximo 4."),
    ("TELÉFONO", "OBLIGATORIO. Un teléfono por pasajero, en el mismo orden que PAX, "
                 "separados con  |  (ej: +54 9 11 4490-7781 | +54 9 11 6033-2210)."),
    ("TIPO", "Llegada (in) = arribo con vuelo · Salida (out) = salida con vuelo · "
             "Hs Disposición = horas a disposición · Otro = traslado."),
    ("ORIGEN / DESTINO", "Dirección o lugar (el sistema usa Google Maps para geolocalizar la dirección)."),
    ("DESTINO 2 / DESTINO 3", "Tramos adicionales del mismo viaje. Dejar vacío si hay un solo tramo."),
    ("VUELO", "Solo para tipo Llegada (in) / Salida (out), ej: AR1234. Dejar vacío en otros casos."),
    ("OBSERVACIONES", "Texto libre opcional."),
]

row = 3
for label, text in instrucciones:
    ws2.cell(row=row, column=1, value=label).font = bold
    ws2.cell(row=row, column=1).alignment = Alignment(vertical="top")
    c = ws2.cell(row=row, column=2, value=text)
    c.font = norm
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[row].height = 30
    row += 1

out = r"C:\caltamirano\logos2\public\plantilla-viajes.xlsx"
wb.save(out)
print("OK", out)
